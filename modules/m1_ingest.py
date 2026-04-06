"""
Módulo 1: Ingestión y Normalización (Argos Generator)
Parsea el Datasheet de Ingeniería (Excel/XLSX) y genera un JSON estandarizado.
"""
import openpyxl
import json
import os
import re
import logging


class DatasheetParser:
    """
    Extrae datos de un Datasheet de Ingeniería (formato Excel híbrido).
    """

    def __init__(self, filepath: str):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Datasheet no encontrado: {filepath}")
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.sheet = self.wb.active
        self.logger = logging.getLogger(__name__)

    def _get_cell_value(self, row: int, col: int) -> str | None:
        """Obtiene el valor de una celda, retorna None si está vacía."""
        val = self.sheet.cell(row=row, column=col).value
        return str(val).strip() if val is not None else None

    def _find_value_by_keyword(self, keyword: str) -> str | None:
        """
        Busca una fila donde la columna B contenga el keyword y retorna el valor de columna C.
        El keyword se busca de forma parcial e insensible a mayúsculas.
        Nota: En este Excel, columna A está vacía, keywords en B, valores en C.
        """
        keyword_lower = keyword.lower().replace(":", "").strip()
        for row_raw in self.sheet.iter_rows(min_col=1, max_col=3, values_only=True):
            row = tuple(row_raw)  # type: ignore
            # Buscar keyword en columnas A o B (por flexibilidad)
            for col_idx in [0, 1]:  # Columnas A y B
                if len(row) > col_idx and row[col_idx]:
                    cell_text = str(row[col_idx]).lower().replace(":", "").strip()
                    if keyword_lower in cell_text or cell_text.startswith(keyword_lower):
                        # Retornar valor de la siguiente columna
                        value_col = col_idx + 1
                        if len(row) > value_col and row[value_col]:
                            return str(row[value_col]).strip()
        return None

    def _extract_header_fields(self) -> dict:
        """
        Extrae campos del encabezado buscando keywords específicas.
        """
        header_data = {}
        keywords_map = {
            "id_gestion": "n° de certificado",
            "tipo_intervencion": "tipo de intervención",  # Nuevo campo clave
            "sku_principal": "sku",
            "marca": "marca",
            "fabrica": "fábrica",
            "direccion_fabrica": "dirección",
        }

        for json_key, keyword in keywords_map.items():
            value = self._find_value_by_keyword(keyword)
            if value:
                # Para id_gestion, extraer solo el número de "CERTIFICADO XXX"
                if json_key == "id_gestion" and "CERTIFICADO" in value.upper():
                    # Extraer el número: "CERTIFICADO 337" -> "337"
                    match = re.search(r'CERTIFICADO\s+(\d+)', value, re.IGNORECASE)
                    if match:
                        value = match.group(1)
                header_data[json_key] = value

        return header_data

    def _extract_models_and_specs(self) -> tuple[list[str], list[str]]:
        """
        Extrae la lista de modelos y especificaciones técnicas.
        Busca filas donde cualquier columna empiece con "MODELO" o "ESPECIFICACIONES".
        """
        modelos = []
        specs = []

        modelos = []
        specs = []

        # Ampliada búsqueda a 15 columnas para cubrir casos como col H/I
        for row_raw in self.sheet.iter_rows(min_col=1, max_col=15, values_only=True):
            row = tuple(row_raw)  # type: ignore
            if not row:
                continue

            # Buscar keywords en toda la fila (hasta donde leímos)
            for col_idx, cell in enumerate(row):
                if not cell:
                    continue
                    
                cell_text = str(cell).strip().upper()

                # Buscar MODELO N:
                if cell_text.startswith("MODELO"):
                    # Valor en la siguiente columna
                    value_col = col_idx + 1
                    if len(row) > value_col and row[value_col]:
                        model_val = str(row[value_col]).strip()
                        if model_val and model_val not in modelos:
                            modelos.append(model_val)

                # Buscar ESPECIFICACIONES (y variantes extendidas):
                keywords_specs = [
                    "ESPECIFICACIONES", "CARACTERISTICAS", "CARACTERÍSTICAS", 
                    "DATOS TÉCNICOS", "DATOS TECNICOS", "TECHNICAL DATA", "SPECS",
                    "DESCRIPCION", "DESCRIPTION", "PRODUCT DESCRIPTION",
                    "RATING", "INPUT", "OUTPUT", "ALIMENTACION"
                ]
                
                # Búsqueda exacta al inicio o contenido
                if any(k in cell_text for k in keywords_specs):
                    value_col = col_idx + 1
                    if len(row) > value_col and row[value_col]:
                        spec_val = str(row[value_col]).strip()
                        # Evitar agregar nombre del campo como valor
                        if spec_val and spec_val.upper() not in keywords_specs and spec_val not in specs:
                            # Guardar solo el valor puro, sin el prefijo
                            specs.append(spec_val)

        return modelos, specs

    def parse(self, id_gestion: str | None = None) -> dict:
        """
        Ejecuta el parsing completo y retorna el JSON estandarizado con soporte multicertificado.
        
        Args:
            id_gestion: ID de gestión. Si no se provee, se extrae del Excel.
        """
        # 1. Extraer campos del header
        header = self._extract_header_fields()
        
        # Usar id_gestion del header si no se proveyó
        if not id_gestion:
            id_gestion = header.get("id_gestion", os.path.splitext(os.path.basename(self.filepath))[0])

        # 2. Extraer modelos y specs (Método Clásico: Fila a Fila, Col A-D)
        modelos, specs = self._extract_models_and_specs()

        # 2b. Escaneo Profundo (Deep Scan): Buscar tablas de modelos en CUALQUIER hoja y columna
        # Esto cubre Caso 5 (Columna H en Hoja 1, o Tabla en Hoja 4)
        try:
            deep_models = self._deep_scan_models()
            if deep_models:
                self.logger.info(f"Modelos detectados por Deep Scan: {len(deep_models)}")
                
                # Política de fusión:
                # Si el método clásico encontró pocos (<3) y el Deep Scan encontró muchos (>3),
                # asumimos que es un formato tabular y priorizamos Deep Scan.
                if len(modelos) < 3 and len(deep_models) > 3:
                     self.logger.info("Usando Deep Scan como fuente principal de modelos.")
                     modelos = deep_models
                else:
                     # Agregar no duplicados
                     for m in deep_models:
                         if m not in modelos:
                             modelos.append(m)
                
                # Limpiar duplicados manteniendo orden
                modelos = list(dict.fromkeys(modelos))
        except Exception as e:
            self.logger.error(f"Error en Deep Scan: {e}")

        # 2c. AI Fallback para Specs: Si no se encontraron specs con método clásico, usar IA
        if not specs or len(specs) == 0:
            self.logger.info("Specs clásicas no encontradas. Intentando extracción con IA...")
            try:
                from modules.ai_helper import AISpecsHelper
                
                # Extraer muestra de texto del Excel para análisis IA
                excel_text = self._get_text_sample(max_rows=30, max_cols=15)
                
                ai_helper = AISpecsHelper()
                ai_result = ai_helper.extract_specs_from_text(excel_text, context="datasheet")
                
                if ai_result and ai_result.get('full_spec'):
                    # Usar la spec completa extraída por IA
                    specs.append(ai_result['full_spec'])
                    self.logger.info(f"✓ Specs extraídas por IA: {ai_result['full_spec']}")
                else:
                    self.logger.warning("IA no pudo extraer specs técnicas")
                    
            except ImportError:
                self.logger.warning("Módulo ai_helper no disponible. Instalar: pip install google-generativeai")
            except ValueError as e:
                self.logger.warning(f"API key de Gemini no configurada: {e}")
            except Exception as e:
                self.logger.error(f"Error en extracción IA de specs: {e}")

        # 3. Determinar Tipo de Producto y Certificados Requeridos
        tipo_intervencion = header.get("tipo_intervencion", "").upper()
        if "JUGUETES" in tipo_intervencion or "FTALATOS" in tipo_intervencion:
            tipo_producto = "JUGUETES"
            certificados_requeridos = [
                {"tipo": "SEGURIDAD_JUGUETES", "archivo": None},
                {"tipo": "FTALATOS", "archivo": None}
            ]
        else:
            # Por defecto asumimos Seguridad Eléctrica
            tipo_producto = "SEGURIDAD_ELECTRICA"
            certificados_requeridos = [
                {"tipo": "SEGURIDAD_ELECTRICA", "archivo": None}
            ]

        # 4. Construir JSON de salida
        result = {
            "id_gestion": id_gestion,
            "tipo_producto": tipo_producto,
            "certificados_requeridos": certificados_requeridos,
            "sku_principal": header.get("sku_principal"),
            "marca": header.get("marca"),
            "fabrica": header.get("fabrica"),
            "direccion_fabrica": header.get("direccion_fabrica"),
            "specs_tecnicas": specs if len(specs) > 1 else (specs[0] if specs else None),
            "modelos_solicitados": modelos,
        }

        return result



    def _get_text_sample(self, max_rows: int = 30, max_cols: int = 15) -> str:
        """
        Extrae una muestra de texto del Excel para análisis con IA.
        
        Args:
            max_rows: Número máximo de filas a extraer
            max_cols: Número máximo de columnas a extraer
        
        Returns:
            String con formato "R{row}C{col}: {value}" para cada celda
        """
        text_parts = []
        
        for r_idx, row in enumerate(self.sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True), 1):
            for c_idx, cell in enumerate(row, 1):
                if cell and str(cell).strip():
                    text_parts.append(f"R{r_idx}C{c_idx}: {cell}")
        
        return "\n".join(text_parts)

    def _deep_scan_models(self):
        """Busca modelos en tablas en CUALQUIER hoja (incluyendo la activa, col H, etc.)."""
        extended_models = []
        try:
            # Reabrir workbook para iterar hojas
            wb = openpyxl.load_workbook(self.filepath, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Buscar encabezado de modelos en primeras 50 filas y hasta columna 15 (O)
                model_col_idx: int = -1
                start_row: int = -1
                
                # Scan headers ampliado
                for r_idx, row_raw in enumerate(sheet.iter_rows(min_row=1, max_row=50, max_col=15, values_only=True), 1):
                    row = tuple(row_raw)  # type: ignore
                    for c_idx, cell in enumerate(row):
                        if cell and isinstance(cell, str):
                            cell_upper = cell.upper()
                            
                            is_header_candidate = ("MODEL" in cell_upper or "CÓDIGO" in cell_upper or "CODIGO" in cell_upper or "ITEM" in cell_upper)
                            is_not_field_label = not cell_upper.endswith(":") 
                            
                            # Evitar "MODELO 1" tipo label
                            is_not_numbered_label = not re.match(r'MODELO\s*\d+$', cell_upper)
                            
                            if is_header_candidate and "MARCA" not in cell_upper and is_not_field_label and is_not_numbered_label:
                                # VALIDACIÓN DE CONTENIDO (Lookahead)
                                # Mirar las siguientes 3 filas para ver si hay datos consistentes
                                valid_candidates_count: int = 0
                                check_rows: int = 5
                                next_start_row: int = int(r_idx) + 1
                                
                                for k in range(check_rows):
                                    val = self._get_cell_value(next_start_row + k, int(c_idx) + 1) # _get_cell_value usa 1-based index
                                    if val and len(val) > 1 and val.upper() not in ["MODELO", "ITEM", "CODIGO"]:
                                        valid_candidates_count += 1
                                
                                # Si al menos 2 de las 5 filas siguientes tienen datos válidos, aceptamos
                                if valid_candidates_count >= 2:
                                    model_col_idx = c_idx
                                    start_row = next_start_row
                                    self.logger.info(f"   [DEEP SCAN] Header válido '{cell}' en {sheet_name}, Fila {r_idx}, Col {c_idx+1}")
                                    break
                                else:
                                    self.logger.debug(f"   [DEEP SCAN] Header descartado '{cell}' (sin datos abajo)")

                    if model_col_idx != -1: break
                
                if model_col_idx != -1:
                    # Extraer modelos y specs de esa columna
                    empty_count = 0
                    current_sheet_models = []
                    
                    # Extraer también specs asociadas (Columna derecha +1)
                    # Heurística: Si la columna derecha tiene datos, asumimos que son specs
                    spec_col_idx = model_col_idx + 1
                    
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, min_col=model_col_idx+1, max_col=spec_col_idx+1, values_only=True), start=start_row):
                        val_model = row[0]
                        val_spec = row[1] if len(row) > 1 else None
                        
                        if val_model:
                            val_str = str(val_model).strip()
                            if len(val_str) > 1 and val_str.upper() not in ["MODELO", "ITEM", "CODIGO"]:
                                current_sheet_models.append(val_str)
                            empty_count = 0
                        else:
                            empty_count += 1
                            if empty_count > 10: 
                                break
                    
                    if len(current_sheet_models) > 0:
                        extended_models.extend(current_sheet_models)
                                
        except Exception as e:
            self.logger.error(f"Error en Deep Scan: {e}")
            
        return extended_models

    def to_json(self, id_gestion: str | None = None, indent: int = 2) -> str:
        """Retorna el resultado como string JSON formateado."""
        data = self.parse(id_gestion)
        return json.dumps(data, ensure_ascii=False, indent=indent)


# --- Función de utilidad para uso rápido ---
def parse_datasheet(filepath: str, id_gestion: str | None = None) -> dict:
    """
    Helper function para parsear un datasheet y obtener el JSON.
    """
    parser = DatasheetParser(filepath)
    return parser.parse(id_gestion)


if __name__ == "__main__":

    # Test rápido con el datasheet de ejemplo
    import sys
    
    test_file = "CAFEXP37 datasheet.xlsx"
    if len(sys.argv) > 1:
        test_file = sys.argv[1]

    print(f"--- Parseando Datasheet: {test_file} ---")
    try:
        parser = DatasheetParser(test_file)
        result = parser.parse("CERTIFICADO 837")
        
        print("\n--- JSON Generado ---")
        print(json.dumps(result, ensure_ascii=False, indent=2))
        
        print(f"\n--- Resumen ---")
        print(f"SKU Principal: {result.get('sku_principal')}")
        print(f"Marca: {result.get('marca')}")
        print(f"Fábrica: {result.get('fabrica')}")
        print(f"Modelos encontrados: {len(result.get('modelos_solicitados', []))}")
        for i, m in enumerate(result.get('modelos_solicitados', []), 1):
            print(f"  {i}. {m}")
            
    except Exception as e:
        print(f"Error: {e}")
